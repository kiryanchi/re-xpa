import sys
import os
import logging
import openpyxl
import openpyxl_image_loader
from PyQt5.QtWidgets import QApplication, QWidget, QFileDialog, QMainWindow, QTabBar, QTabWidget, QScrollArea, QTableWidget, QHeaderView, QAbstractItemView, QDialog
from PyQt5 import uic
from PyQt5.QtCore import pyqtSlot, pyqtSignal
from PyQt5 import QtCore

# logging 설정
logging.basicConfig(
        level = logging.INFO,
        datefmt = '%Y/%m/%d %I:%M:%S %p',
        format = "%(asctime)s [%(levelname)s] - %(message)s "
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
managerUI = uic.loadUiType('UI/manager.ui')[0]
loaddialogUI = uic.loadUiType('UI/loaddialog.ui')[0]

class SheetInnerTableWidget(QTableWidget):
    def __init__(self):
        super().__init__(5, 1)
        self.setAcceptDrops(True)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setAlternatingRowColors(True)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.insertRow(self.rowCount())


class ManagerMainWindow(QMainWindow, managerUI):

    def __init__(self, fileInfo):
        super().__init__()
        logging.info(fileInfo)
        self.setupUi(self)

        self._initWindow(fileInfo)
    
        self.addSheet_pushButton.clicked.connect(self.addSheet_pushButtonClicked)
        self.removeSheet_pushButton.clicked.connect(self.removeSheet_pushButtonClicked)
        self.addCell_pushButton.clicked.connect(self.addCell_pushButtonClicked)
        self.removeCell_pushButton.clicked.connect(self.removeCell_pushButtonClicked)

        self.open_action.triggered.connect(self.open_actiontriggered)
        self.save_action.triggered.connect(self.save_actiontriggered)
        self.saveas_action.triggered.connect(self.saveas_actiontriggered)

        self.undo_action.triggered.connect(self.undo_actiontriggered)
        self.redo_action.triggered.connect(self.redo_actiontriggered)
        self.copy_action.triggered.connect(self.copy_actiontriggered)
        self.paste_action.triggered.connect(self.paste_actiontriggered)
        self.delete_action.triggered.connect(self.delete_actiontriggered)

    def _initWindow(self, fileInfo):
        self.fileName_label.setText(fileInfo['fileName'].split('/')[-1])
        def _gimhaeInit_(sheetNames):
            pass

        def _cjInit_(sheetNames):
            pass

        def _girlInit_(sheetNames):
            pass

        try:
            self.wb = openpyxl.load_workbook(fileInfo['fileName'])
        except FileNotFoundError:
            logging.error('ManagerMainWindow : _initWindow Error')
        else:
            if fileInfo['fileType'] == '김해':
                _gimhaeInit_(self.wb.sheetnames)
            
            elif fileInfo['fileType'] == 'CJ':
                _cjInit_(self.wb.sheetnames)

            elif fileInfo['fileType'] == '여성':
                _girlInit_(self.wb.sheetnames)


    @pyqtSlot()
    def addSheet_pushButtonClicked(self):
        self._addSheet_('새 탭')

    @pyqtSlot()
    def removeSheet_pushButtonClicked(self):
        self._removeSheet_()

    @pyqtSlot()
    def addCell_pushButtonClicked(self):
        self._addCell_()

    @pyqtSlot()
    def removeCell_pushButtonClicked(self):
        self._removeCell_()

    @pyqtSlot()
    def open_actiontriggered(self):
        self._openExcel_()

    def _addSheet_(self, text):
        self.sheetList_tabWidget.addTab(SheetInnerTableWidget(), text + f'{self.sheetList_tabWidget.count()}')

    def _addCell_(self):
        pass

    def _removeSheet_(self):
        self.sheetList_tabWidget.removeTab(self.sheetList_tabWidget.currentIndex())

    def _removeCell_(self):
        pass

    def _openExcel_(self):
        fileName = QFileDialog.getOpenFileName(self, 'xlsx file', './', 'Excel Files(*.xlsx)')
        logging.info(fileName[0])
        self.fileName_label.setText(fileName[0].split('/')[-1])

    @pyqtSlot()
    def save_actiontriggered(self):
        pass

    @pyqtSlot()
    def saveas_actiontriggered(self):
        pass
    
    @pyqtSlot()
    def undo_actiontriggered(self):
        pass

    @pyqtSlot()
    def redo_actiontriggered(self):
        pass

    @pyqtSlot()
    def copy_actiontriggered(self):
        pass
    
    @pyqtSlot()
    def paste_actiontriggered(self):
        pass

    @pyqtSlot()
    def delete_actiontriggered(self):
        pass


class LoadDialog(QDialog, loaddialogUI):
    switch_window = pyqtSignal(dict)
    fileInfo = {}

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.start_pushButton.clicked.connect(self.start_pushButtonClicked)
        self.loadFile_pushButton.clicked.connect(self.loadFile_pushButtonClicked)

    def loadFile_pushButtonClicked(self):
        fileName = QFileDialog.getOpenFileName(self, 'xlsx file', './', 'Excel Files(*.xlsx)')
        self.fileInfo['fileName'] = fileName[0]
        logging.info(fileName[0])
        self.fileName_label.setText(fileName[0].split('/')[-1])

    @pyqtSlot()
    def start_pushButtonClicked(self):
        self.fileInfo['fileType'] = self.fileType_comboBox.currentText()
        self.switch_window.emit(self.fileInfo)



class WindowController:
    """
    창을 변환하는 컨트롤러
    다음 창에 필요한 정보를 넘겨주거나 창을 띄우고 없애는 역할
    """
    def __init__(self):
        self.test = "test"
        pass

    def showLoadDialog(self):
        self.loadDialog = LoadDialog()
        self.loadDialog.switch_window.connect(self.showManagerMainWindow)
        self.loadDialog.show()

    def showManagerMainWindow(self, fileInfo):
        self.managerMainWindow = ManagerMainWindow(fileInfo)
        self.loadDialog.close()
        self.managerMainWindow.show()

#    def showXpaMainWindow(self):
#        self.xpaMainWindow = XpaMainWindow()
#        self.xpaMainWindow.switch_window.connect(self.showWorkSpaceWindow)
#        self.xpaMainWindow.show()
#
#    def showWorkSpaceWindow(self, xpaInfo):
#        self.workSpaceWindow = WorkSpaceWindow(xpaInfo)
#        self.xpaMainWindow.close()
#        self.workSpaceWindow.show()



def main():
    app = QApplication(sys.argv)
    windowController = WindowController()
    windowController.showLoadDialog()
    app.exec_()


if __name__ == '__main__':
    main()
